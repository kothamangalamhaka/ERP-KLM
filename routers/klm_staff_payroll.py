from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
import psycopg2
from psycopg2.extras import DictCursor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta

router = APIRouter()

# DB Connection Helper
def get_db_connection():
    return psycopg2.connect(
        dbname=os.getenv("DB_NAME", "your_db"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASS", "password"),
        host=os.getenv("DB_HOST", "localhost"),
        port=os.getenv("DB_PORT", "5432")
    )

def generate_month_sheet(ws, month_str, db_conn):
    year, month = map(int, month_str.split('-'))
    date_obj = datetime(year, month, 1)
    sheet_title = date_obj.strftime("%b-%y").capitalize()
    ws.title = sheet_title

    # --- Excel Styling Definitions ---
    font_bold_white = Font(bold=True, color="FFFFFF", size=14)
    font_bold = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    fill_black = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fill_yellow = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    fill_orange = PatternFill(start_color="E28743", end_color="E28743", fill_type="solid")
    fill_light_orange = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")
    fill_red = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    fill_maroon = PatternFill(start_color="C2410C", end_color="C2410C", fill_type="solid")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Column Widths (Shifted data to B, C, D)
    ws.column_dimensions['A'].width = 3   # Blank Left Column for margin
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # Header (Merged B1 to D1)
    ws.merge_cells('B1:D1')
    ws['B1'] = f"{date_obj.strftime('%B %Y').upper()}"
    ws['B1'].font = Font(bold=True, color="FACC15", size=22)
    ws['B1'].fill = fill_black
    ws['B1'].alignment = align_center

    ws.append(["", "Particulars", "Amount", "Total"])
    for col_idx in range(2, 5): # Columns B, C, D
        cell = ws.cell(row=2, column=col_idx)
        cell.font = font_bold
        cell.fill = fill_yellow
        cell.alignment = align_center

    with db_conn.cursor(cursor_factory=DictCursor) as cur:
        # Fetch Summary
        cur.execute("SELECT * FROM monthly_payroll_summary WHERE month_year = %s", (month_str,))
        summary = cur.fetchone() or {}

        # 1. Cash Balances
        ws.append(["", "", "", ""])
        row_num = ws.max_row
        ws.merge_cells(f'B{row_num}:D{row_num}')
        ws[f'B{row_num}'] = "Cash Balances"
        ws[f'B{row_num}'].font = font_bold
        ws[f'B{row_num}'].fill = fill_orange
        ws[f'B{row_num}'].alignment = align_center

        cf_cash = float(summary.get('cash_in_hand_cf', 0))
        ajil_cash = float(summary.get('cash_from_ajil', 0))
        total_a = cf_cash + ajil_cash

        ws.append(["", "Cash In hand C/F", cf_cash, ""])
        ws.append(["", "Cash Received From Ajils Account", ajil_cash, ""])
        
        ws.append(["", "A. Total Cash Balances", "", total_a])
        for col_idx in range(2, 5):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = fill_light_orange
            cell.font = font_bold
        ws[f'D{ws.max_row}'].alignment = align_right

        # Fetch Employees
        cur.execute("SELECT * FROM employee_data WHERE (start_date IS NULL OR start_date <= %s) ORDER BY id ASC", (f"{year}-{month:02d}-31",))
        employees = cur.fetchall()

        ws.append(["", "", "", ""])
        row_num = ws.max_row
        ws.merge_cells(f'B{row_num}:D{row_num}')
        ws.cell(row=row_num, column=2, value="Staff Salary").fill = fill_orange
        ws.cell(row=row_num, column=2).font = font_bold
        ws.cell(row=row_num, column=2).alignment = align_center

        total_salary_b = 0
        adjustments = summary.get('staff_adjustments', {})

        for idx, emp in enumerate(employees):
            adj = adjustments.get(str(emp['id']), {})
            
            cur.execute("SELECT SUM(normal_hr) as n, SUM(ot_hr) as o FROM attendance_logs WHERE emp_id = %s AND log_date::text LIKE %s", (emp['id'], f"{month_str}%"))
            logs = cur.fetchone()
            norm_hr = float(logs['n'] or 0)
            ot_hr = float(logs['o'] or 0)
            
            rate = float(emp['base_salary'] or 0) / 260
            basic = norm_hr * rate
            ot = ot_hr * rate
            bonus = float(adj.get('bonus', 0))
            fuel = float(adj.get('fuel', 0))
            mob = float(adj.get('mobile', emp.get('mobile_allowance', 0)))
            remit = float(adj.get('remittance', 0))
            
            staff_tot = basic + ot + bonus + fuel + mob + remit
            total_salary_b += staff_tot

            ws.append(["", f"Staff {idx+1}: {emp['name']}", "", ""])
            ws[f'B{ws.max_row}'].font = font_bold
            
            ws.append(["", f"1. Basic Salary ({norm_hr} Hr)   Rate: {rate:.2f}", basic, ""])
            ws.append(["", f"2. Over Time ({ot_hr} Hr)   Rate: {rate:.2f}", ot, ""])
            ws.append(["", "3. Bonus", bonus, ""])
            ws.append(["", "4. Allowance ::", "", ""])
            ws.append(["", "    a. Fuel Allowance", fuel, ""])
            ws.append(["", "    b. Mobile Allowance", mob, ""])
            ws.append(["", "5. Employee Remittance / Other", remit, ""])
            
            ws.append(["", f"Salary Payable To {emp['name']}", "", staff_tot])
            for col_idx in range(2, 5):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.fill = fill_light_orange
                cell.font = font_bold
            ws[f'D{ws.max_row}'].alignment = align_right

        ws.append(["", "B. Total Salary payable To All Staff", "", total_salary_b])
        for col_idx in range(2, 5):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = fill_light_orange
            cell.font = font_bold
        ws[f'D{ws.max_row}'].alignment = align_right

        # --- Office Expense :: Current Period ---
        import json
        def parse_expenses(data):
            if isinstance(data, str):
                try: return json.loads(data)
                except: return []
            return data or []

        current_expenses = parse_expenses(summary.get('current_expenses', []))
        upcoming_expenses = parse_expenses(summary.get('upcoming_expenses', []))

        ws.append(["", "", "", ""]) # Extra Blank row added here for spacing above Current Period
        ws.append(["", "", "", ""]) # Blank spacing row for the title background (optional, but keeps logic same)
        
        row_num = ws.max_row
        ws.merge_cells(f'B{row_num}:D{row_num}')
        ws[f'B{row_num}'] = "Office Expense :: Current Period"
        ws[f'B{row_num}'].font = font_bold
        ws[f'B{row_num}'].fill = fill_orange
        ws[f'B{row_num}'].alignment = align_center

        total_exp_c = 0
        for exp in current_expenses:
            desc = exp.get('desc', '')
            amt = float(exp.get('amt', 0))
            total_exp_c += amt
            ws.append(["", desc, amt, ""])

        # C. Total Cost of Expenditure
        short_month = date_obj.strftime('%b').upper()
        short_year = str(year)[-2:]
        ws.append(["", f"C. Total Cost of Expenditure {short_month} - {short_year}", "", total_exp_c])
        for col_idx in range(2, 5):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = fill_light_orange
            cell.font = font_bold
        ws[f'D{ws.max_row}'].alignment = align_right

        # Cash Available in Hand
        ws.append(["", "", "", ""]) # Blank spacing row
        cash_available = total_a - (total_salary_b + total_exp_c)
        ws.append(["", "Cash Available in Hand", "", cash_available])
        for col_idx in range(2, 5):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = fill_red
            cell.font = font_bold_white
        ws[f'D{ws.max_row}'].alignment = align_right

        # --- Office Expense :: Upcoming Period ---
        ws.append(["", "", "", ""]) # Blank spacing row
        row_num = ws.max_row
        ws.merge_cells(f'B{row_num}:D{row_num}')
        ws[f'B{row_num}'] = "Office Expense :: Upcoming Period"
        ws[f'B{row_num}'].font = font_bold
        ws[f'B{row_num}'].fill = fill_orange
        ws[f'B{row_num}'].alignment = align_center

        total_exp_d = 0
        for exp in upcoming_expenses:
            desc = exp.get('desc', '')
            amt = float(exp.get('amt', 0))
            total_exp_d += amt
            ws.append(["", desc, amt, ""])

        # D. Total Expected Expenditure
        ws.append(["", "D. Total Expected Expenditure", "", total_exp_d])
        for col_idx in range(2, 5):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = fill_light_orange
            cell.font = font_bold
        ws[f'D{ws.max_row}'].alignment = align_right

        # Expected Cash in Hand
        ws.append(["", "", "", ""]) # Blank spacing row
        expected_cash = cash_available - total_exp_d
        ws.append(["", "Expected Cash in Hand", "", expected_cash])
        for col_idx in range(2, 5):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.fill = fill_maroon
            cell.font = font_bold_white
        ws[f'D{ws.max_row}'].alignment = align_right

    # Final Formatting: Borders, Height, Decimals, Vertical Alignment
    for row in ws.iter_rows(min_row=1, max_col=4, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 30 if row[0].row == 1 else 22
        
        # Check if the row is a blank spacer (no data in B, C, D)
        is_blank_row = all(cell.value in [None, ""] for cell in row[1:])
        
        for cell in row:
            if cell.column > 1: # Column B, C, D
                # Apply borders only if it's not a blank spacer row
                if cell.row > 1 and not is_blank_row:
                    cell.border = thin_border
                
                # Setup decimals for numbers
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
                
                # Center vertically
                if cell.alignment:
                    cell.alignment = Alignment(horizontal=cell.alignment.horizontal, vertical='center')
                else:
                    cell.alignment = Alignment(vertical='center')
                    
    # Freeze the first 2 rows (Title and Header)
    ws.freeze_panes = 'A3'


@router.get("/export/excel")
async def export_excel(type: str, month: str = None, from_month: str = Query(None, alias="from"), to_month: str = Query(None, alias="to"), token: str = None):
    # Verify token if required...
    
    wb = openpyxl.Workbook()
    conn = get_db_connection()
    
    try:
        if type == 'current' and month:
            generate_month_sheet(wb.active, month, conn)
            filename = f"Payroll_Report_{month}.xlsx"
        
        elif type == 'batch' and from_month and to_month:
            start_date = datetime.strptime(from_month, "%Y-%m")
            end_date = datetime.strptime(to_month, "%Y-%m")
            
            current = start_date
            first = True
            while current <= end_date:
                m_str = current.strftime("%Y-%m")
                ws = wb.active if first else wb.create_sheet()
                generate_month_sheet(ws, m_str, conn)
                first = False
                current += relativedelta(months=1)
                
            filename = f"Payroll_Batch_{from_month}_to_{to_month}.xlsx"
        
        else:
            raise HTTPException(status_code=400, detail="Invalid parameters")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"'
        }
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    finally:
        conn.close()
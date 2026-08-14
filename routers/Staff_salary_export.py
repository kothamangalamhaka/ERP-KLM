from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
import pandas as pd
import io
import os
from dotenv import load_dotenv
from datetime import datetime
from dateutil.relativedelta import relativedelta
from sqlalchemy import create_engine, text
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Load environment variables from .env file
load_dotenv()

router = APIRouter()

# Fetch DB credentials from .env
DB_USER = os.getenv("DB_USER")
DB_PASS = os.getenv("DB_PASS")
DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME")

if not all([DB_USER, DB_PASS, DB_NAME]):
    raise ValueError("Database credentials (DB_USER, DB_PASS, DB_NAME) are missing in the .env file")

# Construct the SQLAlchemy Database URL
DATABASE_URL = f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(DATABASE_URL)

@router.get("/export-salary")
def export_salary_to_excel(
    start_period: str = Query(..., description="Format YYYY-MM"),
    end_period: str = Query(..., description="Format YYYY-MM")
):
    try:
        start_date = datetime.strptime(start_period, "%Y-%m")
        end_date = datetime.strptime(end_period, "%Y-%m")
        
        # മാസം തിരിച്ചുള്ള ലിസ്റ്റ് തയ്യാറാക്കുന്നു
        months_list = []
        current = start_date
        while current <= end_date:
            months_list.append(current.strftime("%Y-%m"))
            current += relativedelta(months=1)
            
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for month_year in months_list:
                # Tab Name Format: mmm-yy (Eg: Jan-26)
                date_obj = datetime.strptime(month_year, "%Y-%m")
                tab_name = date_obj.strftime("%b-%y")
                
                # SQL Query - Exact UI Order
                query = text("""
                    WITH active_emps AS (
                        SELECT id, name, designation as current_desig, category
                        FROM employees 
                        WHERE joining_date <= (CAST(:month_year_full AS DATE) + INTERVAL '1 month' - INTERVAL '1 day')
                        AND (category != 'Released' OR released_date >= CAST(:month_year_full AS DATE))
                    ),
                    hist_data AS (
                        SELECT employee_id, previous_designation, previous_category
                        FROM employee_history
                        WHERE start_date <= (CAST(:month_year_full AS DATE) + INTERVAL '1 month' - INTERVAL '1 day')
                        AND end_date >= CAST(:month_year_full AS DATE)
                    ),
                    payroll_data AS (
                        SELECT * FROM staff_payroll WHERE month_year = :month_year_str
                    )
                    SELECT 
                        ROW_NUMBER() OVER (ORDER BY e.name ASC) as "SN",
                        e.name as "Employee Name",
                        COALESCE(h.previous_designation, h.previous_category, e.current_desig, e.category) as "Designation",
                        COALESCE(p.basic_salary, 0) as "Basic Salary",
                        COALESCE(p.over_time, 0) as "Over Time",
                        COALESCE(p.food_allowance, 0) as "Food Allowance",
                        COALESCE(p.mobile_allowance, 0) as "Mobile Allowance",
                        (COALESCE(p.basic_salary, 0) + COALESCE(p.over_time, 0) + COALESCE(p.food_allowance, 0) + COALESCE(p.mobile_allowance, 0)) as "Gross Salary",
                        COALESCE(p.present_days, 0) as "Present Days",
                        COALESCE(p.commission, 0) as "Commission",
                        COALESCE(p.staff_remittance, 0) as "Staff Remittance",
                        COALESCE(p.deduction, 0) as "Deductions",
                        ((COALESCE(p.basic_salary, 0) + COALESCE(p.over_time, 0) + COALESCE(p.food_allowance, 0) + COALESCE(p.mobile_allowance, 0) + COALESCE(p.commission, 0) + COALESCE(p.staff_remittance, 0)) - COALESCE(p.deduction, 0)) as "Total Payable",
                        COALESCE(p.currency, 'SAR') as "Currency",
                        COALESCE(p.status, 'Unpaid') as "Status",
                        p.remark as "Remark"
                    FROM active_emps e
                    LEFT JOIN hist_data h ON e.id = h.employee_id
                    LEFT JOIN payroll_data p ON e.id = p.emp_id
                """)
                
                with engine.connect() as conn:
                    df = pd.read_sql(query, conn, params={
                        "month_year_full": f"{month_year}-01", 
                        "month_year_str": month_year
                    })
                
                if not df.empty:
                    # 🟢 Custom Designation Sorting Logic (Same as UI)
                    def get_designation_order(desig):
                        val = str(desig).lower().replace('.', ' ').replace('-', ' ').replace('_', ' ').replace('/', ' ').strip()
                        if 'managing director' in val or 'director' in val:
                            return 1
                        if 'co ordinator' in val or 'coordinator' in val:
                            return 2
                        if 'jr account' in val or 'junior account' in val:
                            return 4
                        if 'account' in val:
                            return 3
                        if 'office admin' in val:
                            return 5
                        return 999

                    # Apply sorting order
                    df['sort_order'] = df['Designation'].apply(get_designation_order)
                    df = df.sort_values(by=['sort_order', 'Employee Name'])
                    df = df.drop(columns=['sort_order'])
                    
                    # 🟢 Re-assign Serial Number (SN) after sorting
                    df['SN'] = range(1, len(df) + 1)

                # Zero values replacement (Make 0 or Null as blank)
                cols_to_blank = ["Basic Salary", "Over Time", "Food Allowance", "Mobile Allowance", "Gross Salary", "Present Days", "Commission", "Staff Remittance", "Deductions", "Total Payable"]
                for c in cols_to_blank:
                    if c in df.columns:
                        df[c] = df[c].apply(lambda x: "" if pd.isna(x) or x == 0 or x == 0.0 else x)

                # ഡാറ്റ ഇല്ലാത്ത മാസം ആണെങ്കിൽ Empty Columns വരാൻ (Exact Order)
                if df.empty:
                    df = pd.DataFrame(columns=["SN", "Employee Name", "Designation"] + cols_to_blank + ["Currency", "Status", "Remark"])
                    
                df.to_excel(writer, sheet_name=tab_name, index=False)
                
                # ---------------------------------------------------------
                # Excel Styles Configuration (Filter, Black Header, Borders, Align)
                # ---------------------------------------------------------
                worksheet = writer.sheets[tab_name]
                
                header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                center_alignment = Alignment(horizontal="center", vertical="center")
                left_alignment = Alignment(horizontal="left", vertical="center")
                thin_border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                
                # Apply Header Style and AutoFilter
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = thin_border
                worksheet.auto_filter.ref = worksheet.dimensions

                # Apply Borders and Alignments to Data
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                        
                        # Columns 4 to 15 (Basic Salary up to Status) center aligned
                        if 4 <= cell.column <= 15:
                            cell.alignment = center_alignment
                        else:
                            # SN, Name, Designation, Remark left aligned
                            cell.alignment = left_alignment

                # Adjust column widths dynamically for better UI in Excel
                for col in worksheet.columns:
                    col_letter = col[0].column_letter
                    if col_letter == 'A':
                        worksheet.column_dimensions[col_letter].width = 6   # SN
                    elif col_letter == 'B':
                        worksheet.column_dimensions[col_letter].width = 25  # Employee Name
                    elif col_letter == 'C':
                        worksheet.column_dimensions[col_letter].width = 20  # Designation
                    else:
                        worksheet.column_dimensions[col_letter].width = 15  # All other columns

        output.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="Own Staff Salary.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        
        return StreamingResponse(output, headers=headers)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
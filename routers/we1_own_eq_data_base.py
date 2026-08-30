from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime

router = APIRouter()

# DB Connection Helper
def get_db_connection():
    try:
        conn = psycopg2.connect(
            host=os.getenv("DB_HOST", "localhost"),
            database=os.getenv("DB_NAME", "erp_database"),
            user=os.getenv("DB_USER", "postgres"),
            password=os.getenv("DB_PASS", "password"),
            port=os.getenv("DB_PORT", "5432")
        )
        return conn
    except Exception as e:
        print(f"Database connection error: {e}")
        return None

@router.get("/py/api/we1-own-eq/export")
def export_we1_own_eq_excel():
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database connection failed")
        
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        # 🟢 Fetch Master Data with Driver Log (IQAMA & Licence) Details
        query = """
            SELECT 
                m.*,
                d.iqama_no,
                d.iqama_expiry,
                d.licence_expiry
            FROM we1_own_eq_master m
            LEFT JOIN LATERAL (
                SELECT iqama_no, iqama_expiry, licence_expiry
                FROM we1_driver_log
                WHERE UPPER(TRIM(plate_no)) = UPPER(TRIM(m.plate_no))
                ORDER BY COALESCE(join_date, '1970-01-01'::date) DESC, id DESC
                LIMIT 1
            ) d ON true
            ORDER BY m.id ASC
        """
        cursor.execute(query)
        data = cursor.fetchall()
        
        wb = openpyxl.Workbook()
        
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # ==========================================
        # 📑 TAB 1: Classic Summary (We1_Own_EQ)
        # ==========================================
        ws1 = wb.active
        ws1.title = "We1_Own_EQ"

        headers_tab1 = [
            "SN", "Mobilisation Date", "Vehicle Type", "Plate No", "Driver Name", 
            "Mobile", "Joining Date", "Site Name", "Vehicle Owner", "Santhook", 
            "Salary", "Status", "Note"
        ]
        ws1.append(headers_tab1)

        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for i, row in enumerate(data, start=1):
            salary_val = float(row['salary']) if row['salary'] else None
            ws1.append([
                i,
                row['mob_date'],
                row['vehicle_type'] or "",
                row['plate_no'] or "",
                row['driver_name'] or "",
                row['driver_mobile'] or "",
                row['joining_date'],
                row['site_name'] or "",
                row['vehicle_owner'] or "",
                row['santhook'] or "",
                salary_val,
                row['status'] or "Running",
                row['note'] or ""
            ])

        for row in ws1.iter_rows(min_row=2, max_col=13, max_row=len(data) + 1):
            if row[1].value:
                row[1].number_format = 'DD-MMM-YY'
                row[1].alignment = center_alignment
            if row[6].value:
                row[6].number_format = 'DD-MMM-YY'
                row[6].alignment = center_alignment
            if row[12].value:
                row[12].alignment = Alignment(wrap_text=True, vertical="top")

        ws1.auto_filter.ref = f"A1:M{len(data) + 1}"
        ws1.freeze_panes = "A2"
        ws1.sheet_view.zoomScale = 100

        col_widths_tab1 = {
            'A': 6, 'B': 16, 'C': 18, 'D': 14, 'E': 22, 
            'F': 16, 'G': 16, 'H': 22, 'I': 22, 'J': 16, 
            'K': 12, 'L': 12, 'M': 40
        }
        for col, width in col_widths_tab1.items():
            ws1.column_dimensions[col].width = width

        # ==========================================
        # 📑 TAB 2: Full UI Details (Full_Details)
        # ==========================================
        ws2 = wb.create_sheet(title="Full_Details")

        headers_tab2 = [
            "SN", "Mobilisation Date", "Vehicle Type", "Plate No", "Driver Name", 
            "Mobile", "Joining Date", "Site Name", "Vehicle Owner", "Santhook", 
            "Salary", "Status", "Note", "Old EQ", "New EQ", 
            "IQAMA No.", "IQAMA Exp", "Licence Exp", "Chassis No", "Serial No", 
            "EQ Insurance Exp", "FAHS / MVPI Exp", "Operation Card Exp", "Isthimaara Exp"
        ]
        ws2.append(headers_tab2)

        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for i, row in enumerate(data, start=1):
            salary_val = float(row['salary']) if row['salary'] else None
            ws2.append([
                i,
                row['mob_date'],
                row['vehicle_type'] or "",
                row['plate_no'] or "",
                row['driver_name'] or "",
                row['driver_mobile'] or "",
                row['joining_date'],
                row['site_name'] or "",
                row['vehicle_owner'] or "",
                row['santhook'] or "",
                salary_val,
                row['status'] or "Running",
                row['note'] or "",
                row['old_eq'] or "",
                row['new_eq'] or "",
                row['iqama_no'] or "",
                row['iqama_expiry'],
                row['licence_expiry'],
                row['chassis_no'] or "",
                row['serial_no'] or "",
                row['eq_insurance_exp'],
                row['fahs_mvpi_exp'],
                row['op_card_exp'],
                row['isthimaara_exp']
            ])

        date_col_indexes = [2, 7, 17, 18, 21, 22, 23, 24] # 1-based columns for dates
        for row in ws2.iter_rows(min_row=2, max_col=24, max_row=len(data) + 1):
            for col_idx in date_col_indexes:
                cell = row[col_idx - 1]
                if cell.value:
                    cell.number_format = 'DD-MMM-YY'
                    cell.alignment = center_alignment
            if row[12].value: # Note column
                row[12].alignment = Alignment(wrap_text=True, vertical="top")

        ws2.auto_filter.ref = f"A1:X{len(data) + 1}"
        ws2.freeze_panes = "A2"
        ws2.sheet_view.zoomScale = 100

        col_widths_tab2 = {
            'A': 6, 'B': 16, 'C': 18, 'D': 14, 'E': 22, 
            'F': 16, 'G': 16, 'H': 22, 'I': 22, 'J': 16, 
            'K': 12, 'L': 12, 'M': 35, 'N': 14, 'O': 14,
            'P': 16, 'Q': 16, 'R': 16, 'S': 20, 'T': 20,
            'U': 18, 'V': 18, 'W': 18, 'X': 18
        }
        for col, width in col_widths_tab2.items():
            ws2.column_dimensions[col].width = width

        # Save to BytesIO stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers = {
            'Content-Disposition': 'attachment; filename="We1_Own_EQ_Report.xlsx"'
        }
        
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            cursor.close()
            conn.close()


# Driver Payroll Export Endpoints

@router.get("/py/api/we1-eq-driver-payroll/export-screen")
def export_payroll_screen(month_year: str):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database connection failed")
        
    try:
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        query = """
            SELECT p.plate_no, p.driver_name, p.basic_salary, p.over_time, p.deduction, COALESCE(p.advance_paid, 0) as advance_paid, p.remark
            FROM we1_payroll p WHERE p.month_year = %s
            UNION
            SELECT m.plate_no, m.driver_name, 0 as basic_salary, 0 as over_time, 0 as deduction, 0 as advance_paid, '' as remark
            FROM we1_own_eq_master m WHERE NOT EXISTS (
                SELECT 1 FROM we1_payroll p2 WHERE p2.month_year = %s AND p2.plate_no = m.plate_no AND p2.driver_name = m.driver_name
            ) ORDER BY plate_no ASC, driver_name ASC
        """
        cursor.execute(query, (month_year, month_year))
        data = cursor.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        tab_name = datetime.strptime(month_year, "%Y-%m").strftime("%b-%y")
        ws.title = tab_name
        
        apply_payroll_excel_formatting(ws, data)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers = {'Content-Disposition': f'attachment; filename="We1_Own_EQ_Driver_Payroll_{month_year}.xlsx"'}
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            cursor.close()
            conn.close()

@router.get("/py/api/we1-eq-driver-payroll/export-batch")
def export_payroll_batch(start_month: str, end_month: str):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="Database connection failed")
        
    try:
        start_y, start_m = map(int, start_month.split('-'))
        end_y, end_m = map(int, end_month.split('-'))
        months = []
        y, m = start_y, start_m
        while (y < end_y) or (y == end_y and m <= end_m):
            months.append(f"{y}-{m:02d}")
            m += 1
            if m > 12:
                m = 1
                y += 1
                
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        cursor = conn.cursor(cursor_factory=RealDictCursor)
        
        for my in months:
            query = """
                SELECT p.plate_no, p.driver_name, p.basic_salary, p.over_time, p.deduction, COALESCE(p.advance_paid, 0) as advance_paid, p.remark
                FROM we1_payroll p WHERE p.month_year = %s
                UNION
                SELECT m.plate_no, m.driver_name, 0 as basic_salary, 0 as over_time, 0 as deduction, 0 as advance_paid, '' as remark
                FROM we1_own_eq_master m WHERE NOT EXISTS (
                    SELECT 1 FROM we1_payroll p2 WHERE p2.month_year = %s AND p2.plate_no = m.plate_no AND p2.driver_name = m.driver_name
                ) ORDER BY plate_no ASC, driver_name ASC
            """
            cursor.execute(query, (my, my))
            data = cursor.fetchall()
            
            tab_name = datetime.strptime(my, "%Y-%m").strftime("%b-%y")
            ws = wb.create_sheet(title=tab_name)
            
            apply_payroll_excel_formatting(ws, data)
            
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers = {'Content-Disposition': f'attachment; filename="We1_Own_EQ_Driver_Payroll_Batch_{start_month}_to_{end_month}.xlsx"'}
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn:
            cursor.close()
            conn.close()

def apply_payroll_excel_formatting(ws, data):
    headers = ["SN", "Plate No", "Driver Name", "Basic Salary", "Over Time", "Gross Salary", "Deduction", "Total Salary", "Advance / Paid", "Payable", "Status", "Remark"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
    for i, row in enumerate(data, start=1):
        basic = float(row['basic_salary'] or 0)
        ot = float(row['over_time'] or 0)
        gross = basic + ot
        ded = float(row['deduction'] or 0)
        total = gross - ded
        adv = float(row['advance_paid'] or 0)
        payable = total - adv
        
        status = 'Un Paid'
        if total == payable and total != 0:
            status = 'Un Paid'
        elif total > payable and payable > 0:
            status = 'Partially Paid'
        elif payable == 0 and total > 0:
            status = 'Paid'
        elif payable < 0:
            status = 'Advanced'
        elif total == 0:
            status = 'Un Paid'
            
        ws.append([
            i, row['plate_no'], row['driver_name'], 
            basic if basic != 0 else "", 
            ot if ot != 0 else "", 
            gross if gross != 0 else "", 
            ded if ded != 0 else "", 
            total if total != 0 else "", 
            adv if adv != 0 else "", 
            payable if payable != 0 else "", 
            status, 
            row['remark'] or ''
        ])
        
    ws.freeze_panes = "A2"
    ws.sheet_view.zoomScale = 100
    ws.auto_filter.ref = f"A1:L{len(data)+1}"
    
    for row in ws.iter_rows(min_row=2, max_row=len(data)+1, min_col=4, max_col=11):
        for cell in row:
            cell.alignment = center_align

    col_widths = {'A':6, 'B':15, 'C':25, 'D':15, 'E':15, 'F':15, 'G':15, 'H':15, 'I':15, 'J':15, 'K':15, 'L':35}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
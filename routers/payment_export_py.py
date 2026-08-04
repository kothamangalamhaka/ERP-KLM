# new code - routers/payment_export_py.py

import io
from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from typing import List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/py/payment-export", tags=["Payment Excel Export"])

class SheetData(BaseModel):
    sheet_name: str
    aoa: List[List[Any]]
    merges: List[dict]
    is_analysis: bool = False

class ExportRequest(BaseModel):
    file_name: str
    sheets: List[SheetData]

@router.post("/generate-excel")
async def generate_excel(payload: ExportRequest):
    try:
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=11)
        bold_data_font = Font(name="Calibri", size=11, bold=True)
        
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        for sheet_data in payload.sheets:
            ws = wb.create_sheet(title=sheet_data.sheet_name[:31])
            
            # 1. അതിവേഗം ഡാറ്റ വരിവരിയായി ചേർക്കുന്നു (Optimized Append)
            for row in sheet_data.aoa:
                ws.append([val if (val is not None and str(val).strip() != "") else "" for val in row])

            # 2. മെർജ് ചെയ്യുന്നു (Merge Top Headers)
            for m in sheet_data.merges:
                s = m["s"]
                e = m["e"]
                ws.merge_cells(
                    start_row=s["r"] + 1,
                    start_column=s["c"] + 1,
                    end_row=e["r"] + 1,
                    end_column=e["c"] + 1
                )

            # 3. ഹെഡ്ഡർ നിറങ്ങൾ (Exact Colors)
            col_colors = [
                "1E293B", "1E293B", "1E293B", "1E293B", "1E293B", "1E293B",
                "166534", "166534", 
                "1E40AF", "1E40AF", 
                "6B21A8", "6B21A8", 
                "BE123C", "BE123C", "BE123C", "BE123C", "BE123C", 
                "059669", "059669", "059669", "059669", "059669", 
                "B91C1C", "B91C1C", 
                "334155", "334155", "334155", "334155", "334155", "334155", "334155", 
                "1E293B", "1E293B", "1E293B", "1E293B", 
                "166534", "166534"  
            ]
            if sheet_data.is_analysis:
                col_colors.insert(0, "0F2027")

            max_cols = len(col_colors)

            # 4. ഹെഡ്ഡർ സ്റ്റൈൽ (Rows 1 & 2)
            ws.row_dimensions[1].height = 28
            ws.row_dimensions[2].height = 24
            
            for r_idx in [1, 2]:
                for c_idx in range(1, max_cols + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    color_hex = col_colors[c_idx - 1] if c_idx - 1 < len(col_colors) else "1E293B"
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

            # 5. ഡാറ്റാ സ്റ്റൈൽ & കളർ ഹൈലൈറ്റുകൾ (Fast Row-wise Styling)
            offset = 1 if sheet_data.is_analysis else 0
            fill_nr = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
            fill_ot = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
            fill_bill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
            fill_diff_black = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=max_cols):
                for cell in row:
                    c_idx = cell.column
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                    if c_idx in [7 + offset, 8 + offset]:
                        cell.fill = fill_nr
                    elif c_idx in [9 + offset, 10 + offset]:
                        cell.fill = fill_ot
                    elif c_idx in [11 + offset, 12 + offset]:
                        cell.fill = fill_bill
                    elif c_idx in [13 + offset, 14 + offset] and cell.value:
                        cell.fill = fill_diff_black
                        cell.font = bold_data_font

            # 6. AutoFilter, Freeze Panes & Zoom Level
            last_col_letter = get_column_letter(max_cols)
            ws.auto_filter.ref = f"A2:{last_col_letter}{ws.max_row}"
            ws.freeze_panes = "A3"
            ws.sheet_view.zoomScale = 80

            # 7. Column Widths
            for c_idx in range(1, max_cols + 1):
                ws.column_dimensions[get_column_letter(c_idx)].width = 16

        wb.save(output)
        output.seek(0)

        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={payload.file_name}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
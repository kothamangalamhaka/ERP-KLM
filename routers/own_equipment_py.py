import os
import io
import psycopg2
from psycopg2.extras import execute_batch
from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/py/own-equipment", tags=["Own Equipment Tracker"])

class InlineLogEdit(BaseModel):
    equipment_id: int
    year: int
    month: int
    maintenance_cost: float = 0
    basic_salary: float = 0
    overtime: float = 0
    penalty: float = 0
    santook_rent: float = 0
    kafil_comm: float = 0
    owner_comm: float = 0
    investor_comm: float = 0
    debit: float = 0
    pwas: float = 0
    other_expense: float = 0
    op_revenue: float = 0

class BulkLogUpdateRequest(BaseModel):
    logs: List[InlineLogEdit]

@router.post("/bulk-save-logs")
async def bulk_save_logs(payload: BulkLogUpdateRequest):
    try:
        conn = psycopg2.connect(
            dbname=os.getenv("DB_NAME"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASS"),
            host=os.getenv("DB_HOST", "127.0.0.1"),
            port=os.getenv("DB_PORT", "5432")
        )
        cur = conn.cursor()

        query = """
            INSERT INTO equipment_monthly_logs (
                equipment_id, year, month, maintenance_cost, basic_salary, overtime, 
                penalty, santook_rent, kafil_comm, owner_comm, investor_comm, 
                debit, pwas, other_expense, op_revenue
            )
            VALUES (
                %(equipment_id)s, %(year)s, %(month)s, %(maintenance_cost)s, %(basic_salary)s, %(overtime)s,
                %(penalty)s, %(santook_rent)s, %(kafil_comm)s, %(owner_comm)s, %(investor_comm)s,
                %(debit)s, %(pwas)s, %(other_expense)s, %(op_revenue)s
            )
            ON CONFLICT (equipment_id, year, month)
            DO UPDATE SET 
                maintenance_cost = EXCLUDED.maintenance_cost,
                basic_salary = EXCLUDED.basic_salary,
                overtime = EXCLUDED.overtime,
                penalty = EXCLUDED.penalty,
                santook_rent = EXCLUDED.santook_rent,
                kafil_comm = EXCLUDED.kafil_comm,
                owner_comm = EXCLUDED.owner_comm,
                investor_comm = EXCLUDED.investor_comm,
                debit = EXCLUDED.debit,
                pwas = EXCLUDED.pwas,
                other_expense = EXCLUDED.other_expense,
                op_revenue = EXCLUDED.op_revenue;
        """

        log_dicts = [item.dict() for item in payload.logs]
        execute_batch(cur, query, log_dicts, page_size=100)

        conn.commit()
        cur.close()
        conn.close()

        return {"success": True, "message": f"Successfully bulk updated {len(log_dicts)} logs via Python Engine!"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi.responses import Response

class ExportMultiSheetRequest(BaseModel):
    file_name: str = "Equipment_Wise_Details.xlsx"
    equipments: list

@router.post("/export-multi-sheet-excel")
async def export_multi_sheet_excel(payload: ExportMultiSheetRequest):
    try:
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default empty sheet
        
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        total_fill = PatternFill(start_color="EAEFF8", end_color="EAEFF8", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="000000")
        data_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)
        
        thin_border = Border(
            left=Side(style='thin', color='C0C0C0'), right=Side(style='thin', color='C0C0C0'),
            top=Side(style='thin', color='C0C0C0'), bottom=Side(style='thin', color='C0C0C0')
        )

        headers = [
            "Year", "Month", "Maintenance Cost", "Driver (Basic + OT)", "Penalty", 
            "Santook Rent", "Kafil Commission", "Owner Commission", "Investor Commission", 
            "Debit Note", "PWAS", "Other Expenses", "Operational Revenue", "Gain / Loss"
        ]

        for eq in payload.equipments:
            sheet_title = str(eq.get("plate_no", "Sheet"))[:31].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(title=sheet_title)
            
            # Title Row
            title_text = f"Plate No: {eq.get('plate_no')} | Purchase Date: {eq.get('purchase_date', 'N/A')} | Purchase Cost: {eq.get('purchase_cost', 0)} SAR"
            rem_cost = float(eq.get('remaining_purchase_cost', 0) or 0)
            if rem_cost > 0:
                title_text += f" | Remaining Cost: {rem_cost} SAR"
                
            ws.merge_cells("A1:N1")
            title_cell = ws.cell(row=1, column=1, value=title_text)
            title_cell.font = Font(name="Calibri", size=12, bold=True, color="1E40AF")
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 28
            
            # Header Row
            ws.row_dimensions[2].height = 24
            for c_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=2, column=c_idx, value=h_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border

            # Data Rows (All cells horizontally and vertically CENTER aligned)
            rows_data = eq.get("rows", [])
            row_num = 3
            for r in rows_data:
                ws.row_dimensions[row_num].height = 20
                for c_idx, val in enumerate(r, 1):
                    cell = ws.cell(row=row_num, column=c_idx, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    if c_idx in [1, 2]:
                        cell.font = bold_font
                    elif c_idx == 14:  # Gain/Loss color
                        cell.font = Font(name="Calibri", size=11, bold=True, color="FF0000" if float(val or 0) < 0 else "008000")
                row_num += 1

            # Total Row (Center aligned)
            totals_data = eq.get("totals", [])
            if totals_data:
                ws.row_dimensions[row_num].height = 22
                for c_idx, val in enumerate(totals_data, 1):
                    cell = ws.cell(row=row_num, column=c_idx, value=val)
                    cell.font = bold_font
                    cell.fill = total_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    if c_idx == 14:
                        cell.font = Font(name="Calibri", size=11, bold=True, color="FF0000" if float(val or 0) < 0 else "008000")

            # 🟢 AUTO COLUMN WIDTHS (Skip Row 1 Title to prevent Year column from getting too wide)
            for c_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(c_idx)
                if c_idx == 1:
                    ws.column_dimensions[col_letter].width = 10  # Year column fixed small width
                elif c_idx == 2:
                    ws.column_dimensions[col_letter].width = 11  # Month column fixed width
                else:
                    max_len = 0
                    for row_idx in range(2, ws.max_row + 1):
                        cell_val = ws.cell(row=row_idx, column=c_idx).value
                        if cell_val is not None:
                            max_len = max(max_len, len(str(cell_val)))
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 14)
                
            ws.sheet_view.zoomScale = 85

        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={payload.file_name}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
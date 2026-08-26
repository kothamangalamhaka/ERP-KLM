import os
import io
import psycopg2
from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/py/monthly-summary", tags=["Monthly Summary Export"])

@router.get("/export-excel")
async def export_monthly_summary_excel(from_y: int, from_m: int, to_y: int, to_m: int):
    try:
        conn = psycopg2.connect(
            dbname=os.getenv("DB_NAME"), user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASS"), host=os.getenv("DB_HOST", "127.0.0.1"), port=os.getenv("DB_PORT", "5432")
        )
        cur = conn.cursor()
        
        cur.execute("SELECT id, plate_no FROM equipments ORDER BY id ASC")
        equipments = cur.fetchall()
        
        # 🟢 SQL query to fetch data between From and To Date Range
        query = """
            SELECT * FROM equipment_monthly_logs 
            WHERE (year > %s OR (year = %s AND month >= %s))
              AND (year < %s OR (year = %s AND month <= %s))
            ORDER BY year ASC, month ASC
        """
        cur.execute(query, (from_y, from_y, from_m, to_y, to_y, to_m))
            
        logs = cur.fetchall()
        columns = [desc[0] for desc in cur.description]
        logs_dicts = [dict(zip(columns, row)) for row in logs]
        
        cur.close()
        conn.close()

        output = io.BytesIO()
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        month_names = ["", "January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        
        bold_font = Font(name="Calibri", size=11, bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        fill_op_exp = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
        fill_op_rev = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        fill_net_op = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        fill_pl = PatternFill(start_color="E8D0E8", end_color="E8D0E8", fill_type="solid")

        # 🟢 Generate Sheets for the Date Range
        for y in range(from_y, to_y + 1):
            start_month = from_m if y == from_y else 1
            end_month = to_m if y == to_y else 12
            
            for m in range(start_month, end_month + 1):
                month_logs = [l for l in logs_dicts if l['year'] == y and l['month'] == m]
                # Skip empty months in batch exports, except when explicitly exporting a single month
                if not month_logs and not (from_y == to_y and from_m == to_m):
                    continue
                
                ws = wb.create_sheet(title=f"{month_names[m]} {y}")
                
                # Title
                ws.merge_cells("A1:O1")
                title_cell = ws.cell(row=1, column=1, value=f"Monthly Income & Expenditure Statement :: {month_names[m]} {y}")
                title_cell.font = Font(size=14, bold=True)
                title_cell.alignment = center_align
                ws.row_dimensions[1].height = 25
                
                # Header Row 1
                ws.merge_cells("A2:B2"); ws.cell(row=2, column=1, value="EQ").alignment = center_align
                ws.merge_cells("C2:I2"); ws.cell(row=2, column=3, value="Expenditures").alignment = center_align
                ws.merge_cells("J2:K2"); ws.cell(row=2, column=10, value="Income").alignment = center_align
                ws.merge_cells("L2:N2"); ws.cell(row=2, column=12, value="Commissions").alignment = center_align
                ws.cell(row=2, column=15, value="").alignment = center_align
                
                for col in range(1, 16):
                    ws.cell(row=2, column=col).font = bold_font
                    ws.cell(row=2, column=col).border = thin_border
                
                # Header Row 2
                headers = ["SN", "Plate no", "EQ Maintenance Cost", "Net Salary", "Santook Rent", "Debt", "PWAS", "Other", 
                           "Total Operational Expenses", "Operational revenue", "Net OP Revenue", "Kafil", "Owner", "Investor", "Profit or Loss"]
                
                ws.row_dimensions[3].height = 40
                for col_idx, text in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx, value=text)
                    cell.font = bold_font
                    cell.border = thin_border
                    cell.alignment = center_align
                    
                    if col_idx == 9: cell.fill = fill_op_exp
                    elif col_idx == 10: cell.fill = fill_op_rev
                    elif col_idx == 11: cell.fill = fill_net_op
                    elif col_idx == 15: cell.fill = fill_pl
                    
                # Data Rows
                r_idx = 4
                for i, eq in enumerate(equipments, 1):
                    log = next((item for item in month_logs if item["equipment_id"] == eq[0]), {})
                    
                    maint = float(log.get("maintenance_cost", 0) or 0)
                    basic = float(log.get("basic_salary", 0) or 0)
                    ot = float(log.get("overtime", 0) or 0)
                    penalty = float(log.get("penalty", 0) or 0)
                    net_sal = (basic + ot) - penalty
                    rent = float(log.get("santook_rent", 0) or 0)
                    debt = float(log.get("debit", 0) or 0)
                    pwas = float(log.get("pwas", 0) or 0)
                    other = float(log.get("other_expense", 0) or 0)
                    
                    op_rev = float(log.get("op_revenue", 0) or 0)
                    
                    tot_op_exp = maint + net_sal + rent + debt + pwas + other
                    net_op_rev = op_rev - tot_op_exp
                    
                    kafil = op_rev * (float(log.get("kafil_comm", 0) or 0) / 100)
                    owner = float(log.get("owner_comm", 0) or 0)
                    inv = float(log.get("investor_comm", 0) or 0)
                    
                    tot_comm = kafil + owner + inv
                    pl = net_op_rev - tot_comm
                    
                    v = lambda x: "" if x == 0 else round(x, 2)
                    
                    row_data = [i, eq[1], v(maint), v(net_sal), v(rent), v(debt), v(pwas), v(other), 
                                v(tot_op_exp), v(op_rev), v(net_op_rev), v(kafil), v(owner), v(inv), v(pl)]
                    
                    ws.row_dimensions[r_idx].height = 25
                    
                    for c_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=val)
                        cell.border = thin_border
                        cell.alignment = center_align
                        if c_idx == 9: cell.fill = fill_op_exp
                        elif c_idx == 10: cell.fill = fill_op_rev
                        elif c_idx == 11: cell.fill = fill_net_op
                        elif c_idx == 15: 
                            cell.fill = fill_pl
                            if pl < 0: cell.font = Font(color="FF0000")
                    r_idx += 1
                    
                # Formatting Columns width
                equal_width_cols = [2, 4, 6, 7, 8, 12, 13, 14] 
                for col in range(1, 16):
                    if col in equal_width_cols:
                        ws.column_dimensions[get_column_letter(col)].width = 15 
                    elif col == 15: # 🟢 Profit or Loss Column Width
                        ws.column_dimensions[get_column_letter(col)].width = 18 
                    else:
                        ws.column_dimensions[get_column_letter(col)].width = 14

        if len(wb.sheetnames) == 0:
            wb.create_sheet("No Data")
            
        wb.save(output)
        output.seek(0)
        
        if from_y == to_y and from_m == to_m:
            filename = f"Monthly_Statement_{month_names[from_m]}_{from_y}.xlsx"
        else:
            filename = f"Statement_Batch_{month_names[from_m]}_{from_y}_To_{month_names[to_m]}_{to_y}.xlsx"
            
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
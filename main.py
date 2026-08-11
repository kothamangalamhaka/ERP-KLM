import os
import json
import io
import html
import base64
import re
from datetime import datetime
import psycopg2
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


load_dotenv()

app = FastAPI(title="Haka ERP Python Engine")

# ==========================================
# 🚀 REGISTER MODULAR PYTHON ROUTERS
# ==========================================
# new code - main.py
from routers.own_equipment_py import router as own_equipment_router
app.include_router(own_equipment_router)

from routers.payment_py import router as payment_router
app.include_router(payment_router)

# 🟢 അതിവേഗ Excel Export-നായി പുതിയ Python Router ചേർക്കുന്നു
from routers.payment_export_py import router as payment_export_router
app.include_router(payment_export_router)

# Import the new router
from routers import we1_own_eq_data_base

# Include it in your FastAPI app (usually after creating app = FastAPI())
app.include_router(we1_own_eq_data_base.router)

class ExportRequest(BaseModel):
    headers: list
    rows: list
    sheet_name: str = "Master Data"

@app.post("/py/export-excel")
async def generate_styled_excel(data: ExportRequest):
    try:
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = data.sheet_name
        
        # 1. സ്റ്റൈലുകൾ & കളറുകൾ (Styles & Fills)
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Calibri", size=11)
        
        # Replaced = Light Orange, Released = Light Red
        replaced_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        released_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # 🟢 Expiry Colors
        expired_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
        days15_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        days30_fill = PatternFill(start_color="FFFF71", end_color="FFFF71", fill_type="solid")
        white_font = Font(name="Calibri", size=11, color="FFFFFF")
        black_font = Font(name="Calibri", size=11, color="000000")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        # 2. ഹെഡ്ഡർ എഴുതുന്നു (Text Wrap ഒഴിവാക്കി)
        for col_idx, header in enumerate(data.headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header).upper())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = thin_border
            
        ws.row_dimensions[1].height = 28
        
        # 3. 'STATUS' കോളം ഇൻഡക്സ് കണ്ടെത്തുന്നു
        status_col_idx = -1
        for idx, h in enumerate(data.headers):
            if str(h).strip().upper() == "STATUS":
                status_col_idx = idx
                break
        
        # 4. ഡാറ്റാ റോകൾ എഴുതുന്നു (Color ഉൾപ്പെടെ, Text Wrap ഒഴിവാക്കി + Date Parsing for Excel Tree Filter)
        month_map = {
            "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
            "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12
        }

        for row_idx, row_data in enumerate(data.rows, 2):
            row_status = ""
            if status_col_idx != -1 and len(row_data) > status_col_idx:
                row_status = str(row_data[status_col_idx]).strip().lower()
            
            current_row_fill = None
            if row_status == "replaced":
                current_row_fill = replaced_fill
            elif row_status == "released":
                current_row_fill = released_fill

            for col_idx, value in enumerate(row_data, 1):
                header_name = str(data.headers[col_idx-1]).upper()
                is_date_col = any(k in header_name for k in ["DATE", "EXPIRE", "WORK START", "LAST WORKING DAY", "REACHED"])
                
                final_val = ""
                is_parsed_date = False
                parsed_date_obj = None

                if value is not None and str(value).strip() != "":
                    val_str = str(value).strip()
                    val_str = val_str.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
                    val_str = html.unescape(val_str)
                    final_val = val_str

                    # 🟢 SN, RATE, IQAMA NO എന്നീ കോളങ്ങൾ ആണെങ്കിൽ അക്കമാക്കി (Number) മാറ്റുന്നു
                    if header_name == "SN" or "RATE" in header_name or "IQAMA NO" in header_name:
                        try:
                            # ദശാംശം (Decimal) ഉണ്ടെങ്കിൽ float ആയും അല്ലാത്തവ int ആയും മാറ്റുന്നു
                            if "." in val_str:
                                final_val = float(val_str)
                            else:
                                final_val = int(val_str)
                        except Exception:
                            pass

                    # 🟢 ഡേറ്റ് കോളം ആണെങ്കിൽ അതിനെ യഥാർത്ഥ Python Date Object ആക്കി മാറ്റുന്നു
                    if is_date_col:
                        parts = re.split(r'[\/\- \.]', val_str)
                        if len(parts) == 3:
                            try:
                                d = int(parts[0])
                                m_str = parts[1].upper()[:3]
                                m = month_map.get(m_str) or int(parts[1])
                                y = int("20" + parts[2] if len(parts[2]) == 2 else parts[2])
                                if 1 <= d <= 31 and 1 <= m <= 12 and 1900 <= y <= 2100:
                                    parsed_date_obj = datetime(y, m, d).date()
                                    final_val = parsed_date_obj
                                    is_parsed_date = True
                            except Exception:
                                pass
                    
                cell = ws.cell(row=row_idx, column=col_idx, value=final_val)
                cell.font = data_font
                cell.border = thin_border
                
                # യഥാർത്ഥ Date ആണെങ്കിൽ എക്സലിന് DD-MMM-YYYY എന്ന ഫോർമാറ്റ് നൽകുന്നു
                if is_parsed_date:
                    cell.number_format = "dd-mmm-yyyy"
                
                # 🟢 Expiry Alert കളർ കണക്കാക്കുന്നു
                cell_fill = current_row_fill # ഡീഫോൾട്ട് റോ കളർ ഉണ്ടെങ്കിൽ അത് നൽകുക
                cell_font = data_font

                if row_status == "running" and is_parsed_date and any(k in header_name for k in ["IQAMA EXPIRE", "LICENSE EXPIRE", "LICENCE EXPIRE", "EQ INSURAN", "FAHS MVPI"]):
                    today = datetime.now().date()
                    diff_days = (parsed_date_obj - today).days
                    
                    if diff_days < 0:
                        cell_fill = expired_fill
                        cell_font = white_font
                    elif 0 <= diff_days <= 15:
                        cell_fill = days15_fill
                        cell_font = black_font
                    elif 15 < diff_days <= 30:
                        cell_fill = days30_fill
                        cell_font = black_font

                # സെല്ലിൽ കളറും ഫോണ്ടും നൽകുന്നു
                if cell_fill:
                    cell.fill = cell_fill
                cell.font = cell_font
                
                # അലൈൻമെന്റും wrap_text=False ഉം നൽകുന്നു
                if is_date_col or "SN" == header_name:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
                elif "DAYS WORKED" in header_name or "NUMBER" in header_name:
                    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                    
            ws.row_dimensions[row_idx].height = 22
            
        # 5. ഓട്ടോ കോളം വീതി (Max Width 15 Limit)
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            calculated_width = max(max_len + 3, 10)
            ws.column_dimensions[col_letter].width = min(calculated_width, 15)
            
        # 6. ഓട്ടോ ഫിൽട്ടറും സൂം ലെവലും (80%) സെറ്റ് ചെയ്യുന്നു
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.zoomScale = 80
            
        wb.save(output)
        output.seek(0)
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Master_Database_Export.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

class CellEdit(BaseModel):
    dbId: int
    colName: str
    newValue: str

class BulkUpdateRequest(BaseModel):
    edits: list
    username: str

@app.post("/py/bulk-update-cells")
async def py_bulk_update_cells(payload: BulkUpdateRequest):
    try:
        conn = psycopg2.connect(
            dbname=os.getenv("DB_NAME"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASS"),
            host=os.getenv("DB_HOST", "localhost"),
            port=os.getenv("DB_PORT", "5432")
        )
        cur = conn.cursor()
        
        grouped_edits = {}
        for edit in payload.edits:
            db_id = edit.get("dbId") if isinstance(edit, dict) else edit.dbId
            col_name = edit.get("colName") if isinstance(edit, dict) else edit.colName
            new_val = edit.get("newValue") if isinstance(edit, dict) else edit.newValue

            if db_id not in grouped_edits:
                grouped_edits[db_id] = {}
            grouped_edits[db_id][col_name] = new_val

        for db_id, fields in grouped_edits.items():
            json_payload = json.dumps(fields)
            cur.execute(
                """
                UPDATE erp_records 
                SET record_data = record_data || %s::jsonb,
                    plate_number = COALESCE((%s::jsonb->>'PLATE NUMBER'), plate_number),
                    site = COALESCE((%s::jsonb->>'SITE'), site),
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (json_payload, json_payload, json_payload, db_id)
            )
            
        conn.commit()
        cur.close()
        conn.close()
        
        return {"success": True, "message": f"Successfully updated {len(payload.edits)} cells via Python!"}
    except Exception as e:
        return {"success": False, "message": str(e)}

class ImportRequest(BaseModel):
    fileBase64: str
    importMode: str
    username: str

@app.post("/py/import-excel")
async def py_import_excel(payload: ImportRequest):
    try:
        header, encoded = payload.fileBase64.split(",", 1) if "," in payload.fileBase64 else ("", payload.fileBase64)
        file_bytes = base64.b64decode(encoded)
        
        df = pd.read_excel(io.BytesIO(file_bytes))
        df = df.fillna("")
        df = df.astype(str)

        conn = psycopg2.connect(
            dbname=os.getenv("DB_NAME"),
            user=os.getenv("DB_USER"),
            password=os.getenv("DB_PASS"), 
            host=os.getenv("DB_HOST", "127.0.0.1"),
            port=os.getenv("DB_PORT", "5432")
        )
        cur = conn.cursor()

        if payload.importMode == "rewrite":
            cur.execute("UPDATE erp_records SET deleted_at = CURRENT_TIMESTAMP WHERE deleted_at IS NULL")

        records_inserted = 0
        for index, row in df.iterrows():
            row_dict = row.to_dict()
            plate_number = str(row_dict.get("PLATE NUMBER", row_dict.get("PLATE NO", ""))).strip()
            site = str(row_dict.get("SITE", "")).strip()
            sn_val = row_dict.get("SN", index + 1)
            try:
                sn = int(float(sn_val))
            except:
                sn = index + 1

            if plate_number or site:
                cur.execute(
                    """
                    INSERT INTO erp_records (sn, plate_number, site, record_data)
                    VALUES (%s, %s, %s, %s::jsonb)
                    """,
                    (sn, plate_number, site, json.dumps(row_dict))
                )
                records_inserted += 1

        conn.commit()
        cur.close()
        conn.close()

        return {"success": True, "message": f"Successfully imported {records_inserted} records via Python Engine!"}
    except Exception as e:
        return {"success": False, "message": f"Python Import Error: {str(e)}"}